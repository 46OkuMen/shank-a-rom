# Reinsertion script for 46 Okunen Monogatari: The Shinka Ron.

# TODO: Environment message pointers aren't getting adjusted correctly.
# Either way, the spaces in front of the "volcanic currrents" message are unsightly.

# TODO: Crashes to watch out for:
# 1) On moving to next map - happens when changing length of stuff before 0x10555-ish. (Split this block?)
# Looks like the narration right before it is fine to change the length of. Also the thelodus girl dialogue.
# 2) On reading an encyclopedia entry/evolving - happens when changing length of "You evolved too far..." block.
# 3) On fadeout of title graphic - happens when changing length of "Gaia's Heart" environment message.
# 4) On selecting "Evolve" menu option - happens when changing length of battle options and skill names.

# TODO: Missing battle text:
# "You got %d EVO Genes."
# Looks like this is being properly blanked due to overflow, that's good. Just gotta treat the overflow...
# (Does it actually belong in overflow, though??)

# TODO: Moving overflow to the error block/spare block.
    # Done?) Actually figure out where they are
    # Done) In the rom, replace the jp text with equivalent number of spaces
    # 2) Replace all of the error block with equivalent number of spaces
    # 3) Place all the text in the error block (what about control codes???)
    # 4) Rewrite all the pointer values to point to new locations

from __future__ import division
import os
import math
from binascii import unhexlify
from utils import *
from rominfo import file_blocks, file_location, file_length, creature_block, spare_block, pointer_constants

from openpyxl import load_workbook
from shutil import copyfile
from collections import OrderedDict

script_dir = os.path.dirname(__file__)
src_path = os.path.join(script_dir, 'intermediate_roms')
dest_path = os.path.join(script_dir, 'patched_roms')

src_rom_path = os.path.join(src_path, "46 Okunen Monogatari - The Sinkaron (J) A user.FDI")
dest_rom_path = os.path.join(dest_path, "46 Okunen Monogatari - The Sinkaron (J) A user.FDI")

copyfile(src_rom_path, dest_rom_path)

dump_xls = "shinkaron_dump_test.xlsx"
pointer_xls = "shinkaron_pointer_dump.xlsx"

files_to_translate = ['ST1.EXE', 'SINKA.DAT']
# TODO: ST2.EXE has an issue in finding the first block. Check the file_start value...


def get_translations(file, dump_xls):
    # Parse the excel dump and return a dict full of translation tuples.
    trans = OrderedDict() # translations[offset] = (japanese, english)
    
    wb = load_workbook(dump_xls)
    ws = wb.get_sheet_by_name(file)

    total_rows = total_replacements = 0
    overflow_block_lo, overflow_block_hi = spare_block[file]  # Doesn't need translations.

    for row in ws.rows[1:]:  # Skip the first row, it's just labels
        total_rows += 1
        offset = int(row[0].value, 16)

        if (offset >= overflow_block_lo) and (offset <= overflow_block_hi):
            continue

        japanese = row[2].value
        english = row[4].value

        if english:
            total_replacements += 1
        else:
            english = ""

        trans[offset] = (japanese, english)

    return trans


def get_dat_translations(file, dump_xls):
    # TODO: I fixed the offsets for the .dat files, do I still need this as a separate function?
    # I failed to record accurate offsets for the .dat files, and I'm paying the price.
    # But I can load up a list full of (jp, eng) tuples and replace the first instance of each one,
    # and it should be fine.
    trans = []
    wb = load_workbook(dump_xls)
    ws = wb.get_sheet_by_name(file)

    total_rows = total_replacements = 0
    for row in ws.rows[1:]:
        total_rows += 1
        japanese = row[2].value
        english = row[4].value
        if english:
            total_replacements += 1
        else:
            english = ""
        trans.append((japanese, english))
    translation_percent = int(math.floor((total_replacements / total_rows) * 100))
    print file, str(translation_percent) + "% complete"
    return trans


def get_pointers(file, ptr_xls):
    # Parse the pointer excel, calculate differences in pointer values,
    # and return dictionaries of pointers and diffs.
    ptrs = OrderedDict()              # text_offset: pointer_offset

    pointer_wb = load_workbook(ptr_xls)
    pointer_sheet = pointer_wb.get_sheet_by_name("Sheet1") # For now, they are all in the same sheet... kinda ugly.

    for row in pointer_sheet.rows:
        if row[0].value != file:
            continue                          # Access ptrs for the current file only.

        text_offset = int(row[1].value, 16)
        pointer_offset = int(row[2].value, 16)
        if text_offset in ptrs:
            ptrs[text_offset].append(pointer_offset)
        else:
            ptrs[text_offset] = [pointer_offset]

    return ptrs


def get_file_strings(rom_path):
    file_strings = {}
    for file in files_to_translate:
        start = file_location[file]
        length = file_length[file]
        file_strings[file] = file_to_hex_string(rom_path, start, length)
    return file_strings


def get_block_strings(file, rom_path):
    block_strings = []
    for index, block in enumerate(file_blocks[file]):
        lo, hi = block
        block_length = hi - lo
        block_start = file_location[file] + lo

        block_strings.append(file_to_hex_string(rom_path, block_start, block_length))
    return block_strings


def erase_spare_block(file, block_string):
    if file in spare_block:
        #length = len(block_string)
        #block_string = '0'*length       # Or would it be better to do an ascii space '20'?
        block_string = ''
    return block_string


def edit_pointer(file, text_location, diff, file_string):
    # TODO: It's a misnomer now, since it edits all pointers pointing to a given text_location...
    if diff == 0:
        return file_string

    pointer_constant = pointer_constants[file]
    pointer_locations = pointers[text_location]
    #print "This text has", len(pointer_locations), "depending on it"

    patched_file_string = file_string
    for ptr in pointer_locations:
        #print "text is at", hex(text_location), "so edit pointer at", hex(ptr), "with diff", diff

        old_value = text_location - pointer_constant
        old_bytes = pack(old_value)
        old_bytestring = "{:02x}".format(old_bytes[0]) +"{:02x}".format(old_bytes[1])

        location_in_file_string = ptr*2
        rom_bytestring = original_file_strings[file][location_in_file_string:location_in_file_string+4]
        assert old_bytestring == rom_bytestring, 'Pointer bytestring not equal to value in rom'

        #print hex(pointer_location)
        #print "old:", old_value, old_bytes, old_bytestring

        new_value = old_value + diff

        new_bytes = pack(new_value)
        new_bytestring = "{:02x}".format(new_bytes[0]) + "{:02x}".format(new_bytes[1])
        #print "new:", new_value, new_bytes, new_bytestring

        location_in_string = ptr * 2

        old_slice = file_string[location_in_string:]
        new_slice = old_slice.replace(old_bytestring, new_bytestring, 1)
        patched_file_string = patched_file_string.replace(old_slice, new_slice, 1)

    return patched_file_string


def edit_pointers_in_range(file, file_string, (lo, hi), diff):
    #print "lo hi", hex(lo), hex(hi)
    for n in range(lo+1, hi+1):
        try:
            ptr = pointers[n]
            file_string = file_strings[file]
            patched_file_string = edit_pointer(file, n, diff, file_string)
            file_strings[file] = patched_file_string
        except KeyError:
            continue
        #print file_strings[file][0xdd39*2]
        assert original_file_strings[file][0xdd39*2] == file_strings[file][0xdd39*2], 'byte got changed before here'
    return file_strings[file]


def edit_text(file, translations):
    # Replace each jp bytestring with eng bytestrings in the text blocks.
    # Return a new file string.

    creature_block_lo, creature_block_hi = creature_block[file]

    previous_text_offset = file_blocks[file][0][0]
    pointer_diff = 0
    previous_replacement_offset = 0
    block_string = file_blocks[file][0]
    previous_text_block = 0
    current_text_block = 0
    current_block_start,current_block_end = file_blocks[file][0]
    is_overflowing = False

    overflow_bytestrings = OrderedDict()

    for original_location, (jp, eng) in translations.iteritems():
        file_strings[file] = edit_pointers_in_range(file, file_strings[file], (previous_text_offset, original_location), pointer_diff)
        print hex(original_location), pointer_diff
        current_text_block = get_current_block(original_location, file)
        #print "block #:", current_text_block
        if current_text_block != previous_text_block:
            print "Hey, it's a new block!", hex(original_location)
            pointer_diff = 0
            previous_replacement_offset = 0
            is_overflowing = False
            block_string = block_strings[current_text_block]
            current_block_start, current_block_end = file_blocks[file][current_text_block]
        if current_text_block:
            # Does not update if the current_block is 0... since 0 is falsy.
            # Can't do "or if current_text_bllock == 0", since that's "true or false" (always true)
            previous_text_block = current_text_block
        else:
            #print hex(original_location), "is not part of any block"
            # TODO: Consider resetting the block here.
            # Doesn't seem to apply anywhere... really, it's the pointers between blocks you gotta watch out for.
            pass

        previous_text_offset = original_location

        jp_bytestring = sjis_to_hex_string(jp)
        eng_bytestring = ascii_to_hex_string(eng)

        if eng_bytestring:
            new_text_offset = original_location + len(eng_bytestring)//2 + pointer_diff
            # The bytestring is twice as long as the number of bytes.
        else:
            new_text_offset = original_location + len(jp_bytestring)//2 + pointer_diff

        if new_text_offset > current_block_end and not is_overflowing:
            print hex(new_text_offset), "overflows past", hex(current_block_end)
            print eng
            is_overflowing = True
            start_in_block = (original_location - current_block_start)*2
            overflow_bytestring = original_block_strings[current_text_block][start_in_block:]
            # Store the start and end of the overflow bytestring, to make sure all pointers are adjusted in the range.
            overflow_lo, overflow_hi = original_location, current_block_end

            overflow_bytestrings[(overflow_lo, overflow_hi)] = overflow_bytestring
            print "adding new oveflow bytestring", overflow_bytestring

        if eng == "":
            continue

        if is_overflowing:
            print hex(original_location), "is part of an overflow"
            eng = ""
            # TODO: What else should happen when it's overflowing? Should the pointer still get adjusted?


        # Recalculate this in case eng got erased.
        eng_bytestring = ascii_to_hex_string(eng)

        this_string_diff = ((len(eng_bytestring) - len(jp_bytestring)) // 2)   # since 2 chars per byte

        if (original_location >= creature_block_lo) and (original_location <= creature_block_hi):
            if this_string_diff <= 0:
                eng_bytestring += "00"*(this_string_diff*(-1))
                this_string_diff = ((len(eng_bytestring) - len(jp_bytestring)) // 2)
                assert this_string_diff == 0, 'creature string diff not 0'
                # Should be zero unless something went wrong...
            else:
                # Append the 00s to the jp_bytestring so they get replaced - keep the length the same.
                jp_bytestring += "00"*(this_string_diff)
                this_string_diff = ((len(eng_bytestring) - len(jp_bytestring)) // 2)
                assert this_string_diff == 0, 'creature string diff not 0'

        pointer_diff += this_string_diff

        block_string = block_strings[current_text_block]
        try:
            old_slice = block_string[previous_replacement_offset*2:]
            i = old_slice.index(jp_bytestring)//2
        except ValueError:
            previous_replacement_offset = 0
            old_slice = block_string
            i = old_slice.index(jp_bytestring)//2

        new_slice = old_slice.replace(jp_bytestring, eng_bytestring, 1)
        j = block_strings[current_text_block].index(old_slice)
        new_block_string = block_strings[current_text_block].replace(old_slice, new_slice, 1)
        block_strings[current_text_block] = new_block_string

        previous_replacement_offset += i

    patched_file_string = move_overflow(file, file_strings[file], overflow_bytestrings)

    # TODO: Do I need to get patched_file_string rather than file_strings[file]? Doesn't seem like it...
    patched_file_string = pad_text_blocks(file, block_strings, file_strings[file])

    return patched_file_string


def pad_text_blocks(file, block_strings, file_string):
    patched_file_string = file_string
    for i, blk in enumerate(block_strings):
        block_diff = len(blk) - len(original_block_strings[i])   # if block is too short, negative; too long, positive
        #print len(original_block_strings[i]), len(blk)
        print "padding block #", i
        print block_diff
        assert block_diff <= 0, 'Block ending in %s is too long' % hex(file_blocks[file][i][1])
        if block_diff < 0:
            number_of_spaces = ((-1)*block_diff)//2
            inserted_spaces_index = file_blocks[file][i][1]
            blk += '20' * number_of_spaces       # Fill it up with ascii 20 (space)
            print number_of_spaces, "added at", hex(inserted_spaces_index)

        j = original_file_strings[file].index(original_block_strings[i])

        #print original_block_strings[i]

        #assert original_file_strings[file][0xdd39*2] == file_strings[file][0xdd39*2], 'byte got changed before here'
        j = patched_file_string.index(original_block_strings[i])
        patched_file_string = patched_file_string.replace(original_block_strings[i], blk, 1)

    return patched_file_string


def move_overflow(file, file_string, overflow_bytestrings):
    spare_block_lo, spare_block_hi = spare_block[file]
    #spare_block_string = erase_spare_block(file, block_strings)
    spare_block_string = ''
    location_in_spare_block = 0
    # TODO: Find out why there are duplicate bytestrings being moved.
    for (lo, hi), bytestring in overflow_bytestrings.iteritems():
        # The first pointer must be adjusted to point to the beginning of the spare block.
        #last_pointer_adjustment = lo-1
        pointer_diff = (spare_block_lo - lo)
        # Find all the translations that need to be applied.
        t = OrderedDict()
        for i in range(lo, hi):
            previous_text_location = lo
            if i in translations:
                print "translating overflow"
                print hex(i), pointer_diff
                t[i] = translations[i]
                jp = t[i][0]
                eng = t[i][1]

                jp_bytestring = sjis_to_hex_string(jp)
                eng_bytestring = ascii_to_hex_string(eng)

                this_string_diff = ((len(eng_bytestring) - len(jp_bytestring)) // 2)

                #print jp_bytestring
                #print eng_bytestring
                #print this_string_diff

                j = bytestring.index(jp_bytestring)
                bytestring = bytestring.replace(jp_bytestring, eng_bytestring)
                print "editing pointers between", hex(previous_text_location), hex(i)
                edit_pointers_in_range(file, file_string, (previous_text_location, i), pointer_diff)

                spare_block_string += bytestring

                #print spare_block_string

                previous_text_location = i
                pointer_diff += this_string_diff

    block_strings[-1] = spare_block_string
    original_block_string = original_block_strings[-1]
    file_string = file_string.replace(original_block_string, spare_block_string)
    return file_string


def edit_dat_text(file, file_string):
    translations = get_dat_translations(file, dump_xls)

    patched_file_string = file_string

    for (jp, eng) in translations:
        if eng == "":
            continue
        jp_bytestring = sjis_to_hex_string(jp)
        eng_bytestring = ascii_to_hex_string(eng)

        patched_file_string = patched_file_string.replace(jp_bytestring, eng_bytestring, 1)
    return patched_file_string


def change_starting_map(map_number):
    starting_map_number_location = 0xedaa + file_location['ST1.EXE']
    new_map_bytes = str(map_number).encode()
    f = open(dest_rom_path, 'rb+')
    f.seek(starting_map_number_location)
    f.write(new_map_bytes)
    f.close()

full_rom_string = file_to_hex_string(src_rom_path)
file_strings = get_file_strings(src_rom_path)
original_file_strings = file_strings.copy()

for file in files_to_translate:
    if file == "SINKA.DAT" or file == 'SEND.DAT':
        # Edit the file separately. That'll have to do for now.
        dat_path = os.path.join(src_path, file)
        dest_dat_path = os.path.join(dest_path, file)
        dat_file_string = file_to_hex_string(dat_path)

        patched_dat_file_string = edit_dat_text(file, dat_file_string)

        with open(dest_dat_path, "wb") as output_file:
            data = unhexlify(patched_dat_file_string)
            output_file.write(data)
        continue

    translations = get_translations(file, dump_xls)
    pointers = get_pointers(file, pointer_xls)

    # Then get individual strings of each text block, and put them in a list.
    block_strings = get_block_strings(file, dest_rom_path)
    original_block_strings = list(block_strings)   # Needs to be copied - simple assignment would just pass the ref.

    patched_file_string = edit_text(file, translations)

    i = full_rom_string.index(original_file_strings[file])
    full_rom_string = full_rom_string.replace(original_file_strings[file], patched_file_string, 1)

    # Write the data to the patched file.
    with open(dest_rom_path, "wb") as output_file:
        data = unhexlify(full_rom_string)
        output_file.write(data)

    # Write the translated file alone to the file too.
    dest_file_path = os.path.join(dest_path, file)
    with open(dest_file_path, "wb") as output_file:
        data = unhexlify(patched_file_string)
        output_file.write(data)

    # Get some quick stats on reinsertion progress.
    translated_strings = 0
    total_strings = len(translations)
    for _, eng in translations.itervalues():
        if eng:
            translated_strings += 1

    translation_percent = int(math.floor((translated_strings / total_strings) * 100))
    print file, str(translation_percent), "% complete"

change_starting_map(101)

# 100: open water, volcano cutscene immediately, combat
# 101: caves, hidden hemicyclapsis, Gaia's Heart in upper right
# 102: OOB cladoselache cave
# 103: OOB ???
# 104: OOB Gaia portal
# 105: (default) thelodus sea
# 200: chapter 2 world map
# 201: super glitchy salamander cave
# 202: useful! take one step, die from fresh water, respawn as an Ichthyostega!
# goes until 209.
# 300: black screen. It's on a different disk, of course...


# Useful tip: "Load File 1" takes you to map 105 from any map!